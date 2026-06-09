"""
Report Generation Module
Generates PDF and Excel reports
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, PageBreak, Image
    )
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ReportGenerator:
    """Generate professional reports in PDF and Excel formats"""

    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

    def generate_pdf_report(self, analysis_result: Dict) -> str:
        """Generate a professional PDF report"""
        if not PDF_AVAILABLE:
            return "PDF generation not available. Install reportlab."

        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"hydraulic_analysis_{timestamp}.pdf"
            filepath = self.output_dir / filename

            doc = SimpleDocTemplate(
                str(filepath),
                pagesize=letter,
                rightMargin=0.75*inch,
                leftMargin=0.75*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )

            elements = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1f4788'),
                spaceAfter=6,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            elements.append(Paragraph(
                "HYDRAULIC DRILL MACHINE<br/>PERFORMANCE ANALYSIS REPORT",
                title_style
            ))
            elements.append(Spacer(1, 0.2*inch))

            # Executive Summary
            elements.append(Paragraph("EXECUTIVE SUMMARY", styles['Heading2']))
            summary_data = [
                ["Parameter", "Value", "Status"],
                ["Drill Type", analysis_result.get('drill_type'), ""],
                ["Efficiency", f"{analysis_result.get('efficiency_percentage')}%",
                 analysis_result.get('efficiency_category')],
                ["Machine Health", f"{analysis_result.get('health_score')}%",
                 analysis_result.get('health_status')],
                ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""]
            ]

            summary_table = Table(summary_data, colWidths=[2*inch, 2*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 0.3*inch))

            # Performance Metrics
            elements.append(Paragraph("PERFORMANCE METRICS", styles['Heading2']))
            params = analysis_result.get('input_parameters', {})
            metrics_data = [
                ["Metric", "Value", "Unit"],
                ["Hydraulic Pressure", f"{params.get('pressure', 0)}", "bar"],
                ["Flow Rate", f"{params.get('flow_rate', 0)}", "L/min"],
                ["Rotation Speed", f"{params.get('rpm', 0)}", "RPM"],
                ["Oil Temperature", f"{params.get('temperature', 0)}", "°C"],
                ["Piston Diameter", f"{params.get('piston_diameter', 0)}", "mm"],
                ["Drill Bit Diameter", f"{params.get('bit_diameter', 0)}", "mm"],
            ]

            metrics_table = Table(metrics_data, colWidths=[2.5*inch, 1.5*inch, 1*inch])
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(metrics_table)
            elements.append(Spacer(1, 0.3*inch))

            # Calculated Values
            elements.append(Paragraph("CALCULATED ENGINEERING VALUES", styles['Heading2']))
            calculated_data = [
                ["Parameter", "Value", "Unit"],
                ["Hydraulic Power", f"{analysis_result.get('hydraulic_power_kw', 0)}", "kW"],
                ["Piston Area", f"{analysis_result.get('piston_area_mm2', 0)}", "mm²"],
                ["Hydraulic Force", f"{analysis_result.get('hydraulic_force_n', 0):,.0f}", "N"],
                ["Efficiency Index", f"{analysis_result.get('efficiency_percentage', 0)}", "%"],
            ]

            calc_table = Table(calculated_data, colWidths=[2.5*inch, 1.5*inch, 1*inch])
            calc_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0f0f0')),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(calc_table)
            elements.append(Spacer(1, 0.3*inch))

            # Diagnostics
            diagnostics = analysis_result.get('diagnostics', [])
            if diagnostics:
                elements.append(PageBreak())
                elements.append(Paragraph("DIAGNOSTICS & ALERTS", styles['Heading2']))

                for diagnostic in diagnostics[:10]:  # First 10 diagnostics
                    alert_type = diagnostic.get('type', 'INFO')
                    param = diagnostic.get('parameter', '')
                    msg = diagnostic.get('message', '')
                    rec = diagnostic.get('recommendation', '')

                    elements.append(Paragraph(
                        f"<b>[{alert_type}] {param}</b>",
                        ParagraphStyle('Diagnostic', parent=styles['Normal'],
                                     textColor=colors.red if alert_type == 'CRITICAL' else colors.orange)
                    ))
                    elements.append(Paragraph(msg, styles['Normal']))
                    elements.append(Paragraph(f"<b>Recommendation:</b> {rec}", styles['Normal']))
                    elements.append(Spacer(1, 0.1*inch))

            doc.build(elements)
            return str(filepath)

        except Exception as e:
            return f"Error generating PDF: {str(e)}"

    def generate_excel_report(self, analysis_result: Dict) -> str:
        """Generate a professional Excel report"""
        if not EXCEL_AVAILABLE:
            return "Excel generation not available. Install openpyxl."

        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"hydraulic_analysis_{timestamp}.xlsx"
            filepath = self.output_dir / filename

            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis Report"

            # Styling
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Title
            ws.merge_cells('A1:D1')
            title_cell = ws['A1']
            title_cell.value = "HYDRAULIC DRILL MACHINE PERFORMANCE ANALYSIS"
            title_cell.font = Font(bold=True, size=14, color="1F4788")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 25

            ws.append([])  # Empty row

            # Executive Summary
            ws.append(["EXECUTIVE SUMMARY"])
            ws['A3'].font = Font(bold=True, size=11)

            ws.append(["Drill Type", analysis_result.get('drill_type')])
            ws.append(["Efficiency (%)", analysis_result.get('efficiency_percentage')])
            ws.append(["Efficiency Category", analysis_result.get('efficiency_category')])
            ws.append(["Machine Health (%)", analysis_result.get('health_score')])
            ws.append(["Health Status", analysis_result.get('health_status')])
            ws.append(["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

            ws.append([])  # Empty row

            # Input Parameters
            ws.append(["INPUT PARAMETERS"])
            current_row = ws.max_row
            ws[f'A{current_row}'].font = Font(bold=True, size=11)

            params = analysis_result.get('input_parameters', {})
            ws.append(["Hydraulic Pressure (bar)", params.get('pressure', 0)])
            ws.append(["Flow Rate (L/min)", params.get('flow_rate', 0)])
            ws.append(["Rotation Speed (RPM)", params.get('rpm', 0)])
            ws.append(["Oil Temperature (°C)", params.get('temperature', 0)])
            ws.append(["Piston Diameter (mm)", params.get('piston_diameter', 0)])
            ws.append(["Drill Bit Diameter (mm)", params.get('bit_diameter', 0)])

            ws.append([])  # Empty row

            # Calculated Values
            ws.append(["CALCULATED ENGINEERING VALUES"])
            current_row = ws.max_row
            ws[f'A{current_row}'].font = Font(bold=True, size=11)

            ws.append(["Hydraulic Power (kW)", analysis_result.get('hydraulic_power_kw', 0)])
            ws.append(["Piston Area (mm²)", analysis_result.get('piston_area_mm2', 0)])
            ws.append(["Hydraulic Force (N)", analysis_result.get('hydraulic_force_n', 0)])

            ws.append([])  # Empty row

            # Diagnostics
            diagnostics = analysis_result.get('diagnostics', [])
            if diagnostics:
                ws.append(["DIAGNOSTICS"])
                current_row = ws.max_row
                ws[f'A{current_row}'].font = Font(bold=True, size=11)

                ws.append(["Type", "Parameter", "Message", "Recommendation"])
                current_row = ws.max_row
                for cell in ws[current_row]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                for diagnostic in diagnostics[:20]:  # First 20 diagnostics
                    ws.append([
                        diagnostic.get('type', ''),
                        diagnostic.get('parameter', ''),
                        diagnostic.get('message', ''),
                        diagnostic.get('recommendation', '')
                    ])

            # Adjust column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 35

            wb.save(filepath)
            return str(filepath)

        except Exception as e:
            return f"Error generating Excel: {str(e)}"
